import openpyxl
import getpass
from os import path, system, name
import os
import sys      #for hash
import hashlib  #for hash
import socket

class login:
    def __init__(self):
        self.done = False
        f = open("information/credentials.xlsx",)       #opening needed excel file
        self.path = path.realpath(f.name)               #location of file
        self.wb_obj = openpyxl.load_workbook(self.path) #create workbook object
        self.ws = self.wb_obj.active
        self.rows = self.ws.max_row
        self.username = ''
        
    def UI(self):
        print("welcome...")
        self.done = False
        while self.done == False:
            welcome = input("Do you have an account? Y/N: ")
            if welcome == 'n' or welcome == 'N':
                self.newuser()
                self.newpassword()
            elif welcome == 'y' or welcome == "Y":
                self.username_check()

    def hash(self,passwd):
        code = passwd.encode()
        sha = hashlib.sha512(code)
        return(sha.hexdigest())
        
    def username_check(self):
        username = input("Enter your username: ")
        found = False
        for i in range(1,self.rows):
            usercheck = self.ws.cell(row=i+1, column = 2).value #check if username is taken
            if usercheck == username:
                self.index = i
                self.tempname = username
                found = True
                self.pass_check()
        if found == False:
            print("Username does not exist!")

    def pass_check(self):
        passwd = getpass.getpass(prompt='Password: ', stream=None)
        passcheck = self.ws.cell(row = self.index+1, column = 3).value
        if self.hash(passwd) == passcheck:
            self.username = self.tempname
            self.done = True
        else:
            print("Incorrect Password!")
            leave = input("Enter '0' to return to initial screen, or press any other key to try password again: ")
            if leave == '0':
                self.UI()
            else:
                self.pass_check()
            
    def newuser(self):
        username = input("Enter your username: ")
        for i in range(1,self.rows):
            usercheck = self.ws.cell(row = i+1, column = 2).value
            if usercheck == username:
                print("Username has been taken")
                self.done = False
                break
        else:
            self.username = username
            
    def newpassword(self):
        passwd = getpass.getpass(prompt='Password: ', stream=None)
        passreal = self.hash(passwd)
        self.ws.append([self.rows,self.username,passreal])
        self.wb_obj.save("information/credentials.xlsx")
        self.done = True

b=login()

