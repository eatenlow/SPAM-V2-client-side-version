import openpyxl
import datetime
from os import path, system, name
import socket


class server():
    def __init__(self):
        self.serversocket = socket.socket(socket.AF_INET,socket.SOCK_STREAM)
        self.serversocket.bind(('0.0.0.0',8089))
        self.serversocket.listen(5)# maximum 5 connections, handle one at a time
        self.con,address = self.serversocket.accept()
        self.now = datetime.datetime.now()
        self.today = self.now.strftime('%w')         #get the day of the week in an integer type
        self.todaydate = datetime.date.today()       #get the date today in DD/MM/YYYY format
        self.date = self.todaydate
        self.daydic = {1:'monday.xlsx', 2:'tuesday.xlsx', 3:'wednesday.xlsx',
                4:'thursday.xlsx', 5:'friday.xlsx', 6:'saturday.xlsx', 0:'sunday.xlsx'}

    def userInterface(self):
        print("server running")
        self.receiver()
        while True:
            if self.messageReceived == '1' or self.messageReceived == '2' or self.messageReceived == '3':
                if self.messageReceived == '1':
                    print("show current day menu")
                elif self.messageReceived == '2':
                    print("Searching for item")
                elif self.messageReceived == '3':
                    print("Displaying cart")
                self.MenuRead(self.today)
                self.sender(self.menu)
                print('menu sent')
            elif self.messageReceived == '5':
                print('user in pre order mode')
                self.receiver()
                self.MenuRead(int(self.messageReceived)-1)
                self.sender(self.date)
                self.sender(self.menu)
                
            elif self.messageReceived == 'Checkout':
                print("client checking out")
                self.MenuRead(self.today)
                self.sender(self.menu)
                self.receiver()
                if self.messageReceived == 'preorder':
                    self.preorder()
                    print("preorder orders recorded")
                    self.receiver()
                elif self.messageReceived == 'bill':
                    self.createBill()
                    print("transaction logged")
            elif self.messageReceived == 'DC':
                self.con.close()
                print('Client disconnected')
            self.receiver()
        self.serversocket.close()

    def check_preorder(self):
        print("Checking pre orders logs")
        f = open("information/preorders.xlsx")
        self.path = path.realpath(f.name)
        wb_obj = openpyxl.load_workbook(self.path)
        ws = wb_obj.active
        rows = ws.max_row
        for i in range(1,rows):
            date = ws.cell(row = i+1,column = 1)
            user = ws.cell(row = i+1, column = 4)
            if date.value == self.todaydate and user.value == self.username:
                food = ws.cell(row = i+1,column = 2)
                quantity = ws.cell(row = i+1,column = 3)
                self.sender(str(food.value)+":"+str(quantity.value))
        self.sender("complete")
        
    def preorder(self):
        print('recording pre orders')
        f = open("information/preorders.xlsx")
        self.path = path.realpath(f.name)
        wb_obj = openpyxl.load_workbook(self.path)
        ws = wb_obj.active
        self.receiver()
        temp = self.messageReceived.split(',')
        for i in temp:
            a = i.split(':')
            print(self.date.weekday())
            for single_data in (self.todaydate + datetime.timedelta(n) for n in range (6)):
                print(a[1][-1])
                if self.todaydate.weekday == a[1][-1]:
                    ws.append(self.todaydate,a[0],(a[1][0:len(a[1])-1]),self.username)
        wb_obj.save("information/preorders.xlsx")
        print('preorders recorded')
        

    def createBill(self):
        f = open("information/transactions.xlsx")
        self.path = path.realpath(f.name)
        wb_obj = openpyxl.load_workbook(self.path)
        ws = wb_obj.active
        rows = ws.max_row
        self.time = self.now.strftime("%H:%M:%S")    #get the time in HH/MM/SS format
        self.receiver()
        self.cart = self.messageReceived.split(',')
        for i in self.cart:
            temp = i.split(':')
            ws.append([self.todaydate,self.time,temp[0],temp[1],self.username])
        wb_obj.save("information/transactions.xlsx")

    def userCheck(self):
        if self.username == "Admin":
            self.discount = '0.8'
            print("admin user is logged in")
        else:
            self.discount = '1'
        self.sender(self.discount)
        print("start up process complete")
                       
    def startup(self):
        print("startup")
        self.username = self.con.recv(128).decode()
        self.userCheck()
        self.sender(str(self.todaydate))
        print('date sent')
        self.check_preorder()
    
    def receiver(self):
        print('receiver running')
        reply = self.con.recv(1000)
        print('input received')
        self.messageReceived = reply.decode()
                       
    def sender(self,msg):
        reply = str(msg).encode()
        self.con.sendall(reply)


    def MenuRead(self,day):
        print('reading menu')
        self.menu=''
        choice = self.daydic[int(day)]
        f = open("Menu Excel/"+choice)
        self.path = path.realpath(f.name)           #location of file
        wb_obj = openpyxl.load_workbook(self.path)  #create workbook object of menu
        ws = wb_obj.active
        rows = ws.max_row
        for i in range(1,rows):                     #add food items and prices of the day into a dictionary
            food = ws.cell(row = i+1, column = 1)
            price = ws.cell(row = i+1, column = 2)
            if food.value != '' and price.value != '':
                self.menu += (str(food.value)+":"+str(price.value)+",")
        self.menu = self.menu[0:len(self.menu)-1]
        self.date = self.todaydate                  #used to "reset" the date variable
        while self.date.weekday() != int(day):
            self.date += datetime.timedelta(1)
                       

a=server()
a.startup()
a.userInterface()
